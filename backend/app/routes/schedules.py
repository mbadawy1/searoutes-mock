import csv
import io
import re
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..providers.base import Page, ScheduleFilter, ScheduleProvider

# Import exception classes for error handling
try:
    from ..providers.searoutes import (
        SearoutesAPIError,
        SearoutesError,
        SearoutesRateLimitError,
    )
except ImportError:
    # Fallback in case Searoutes provider is not available
    class SearoutesError(Exception):
        def to_dict(self):
            return {"code": "SEAROUTES_ERROR", "message": str(self)}

    SearoutesRateLimitError = SearoutesAPIError = SearoutesError

router = APIRouter()
provider: ScheduleProvider = None  # Will be injected by main.py


def set_provider(schedule_provider: ScheduleProvider):
    """Inject the schedule provider instance."""
    global provider
    provider = schedule_provider


def normalize_locode_scac(value: Optional[str]) -> Optional[str]:
    """Normalize LOCODE/SCAC: uppercase, strip spaces/hyphens."""
    if not value:
        return value
    # Strip spaces and hyphens, then uppercase
    return re.sub(r"[\s\-]", "", value).upper()


def validate_iso_date(date_str: Optional[str], field_name: str) -> Optional[str]:
    """Validate that date string is ISO-8601 format and return it, or raise HTTPException."""
    if not date_str:
        return date_str

    try:
        # Try to parse as ISO-8601 date
        datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        return date_str
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid date format for '{field_name}'. Expected ISO-8601 format (e.g., '2025-08-20' or '2025-08-20T12:00:00').",
        )


def extract_locode(location: str) -> str:
    """
    Extract LOCODE from location string.
    Maps full location names to UN/LOCODEs for CSV export.
    """
    locode_map = {
        "Alexandria, EG": "EGALY",
        "Tanger, MA": "MATNG",
        "Rotterdam, NL": "NLRTM",
        "Damietta, EG": "EGDAM",
        "Valencia, ES": "ESVLC",
        "Port Said, EG": "EGPSD",
        "Barcelona, ES": "ESBCN",
        "Hamburg, DE": "DEHAM",
        "Le Havre, FR": "FRLEH",
        "Genoa, IT": "ITGOA",
        "Piraeus, GR": "GRPIR",
        "Antwerp, BE": "BEANR",
        "Singapore, SG": "SGSIN",
        "Dubai, AE": "AEDXB",
        "Jeddah, SA": "SAJED",
        # Add more mappings as needed
    }

    return locode_map.get(location, location.split(",")[0][:5].upper())


@router.get("/api/schedules")
def list_schedules(
    origin: Optional[str] = None,
    destination: Optional[str] = None,
    date_from: Optional[str] = Query(None, alias="from"),
    date_to: Optional[str] = Query(None, alias="to"),
    equipment: Optional[str] = None,
    routingType: Optional[str] = None,
    carrier: Optional[str] = None,
    sort: str = "etd",
    page: int = 1,
    pageSize: int = 50,
    nContainers: int = 1,
):
    """
    List schedules with filtering, sorting, and pagination.

    Query parameters:
    - origin: Origin port (partial match)
    - destination: Destination port (partial match)
    - from: Start date for ETD window (alias for date_from)
    - to: End date for ETD window (alias for date_to)
    - equipment: Container equipment type
    - routingType: Direct or Transshipment
    - carrier: Shipping line/carrier name (partial match)
    - sort: Sort field (etd|transit)
    - page: Page number (1-based)
    - pageSize: Items per page
    - nContainers: Number of containers (default 1, affects CO₂ calculations)

    Returns:
    - items: List of schedule objects
    - total: Total number of items matching filters
    - page: Current page number
    - pageSize: Items per page
    """
    # Input hygiene & validation (Task 22)
    # Validate ISO-8601 dates
    date_from = validate_iso_date(date_from, "from")
    date_to = validate_iso_date(date_to, "to")

    # Normalize LOCODE/SCAC inputs
    origin = normalize_locode_scac(origin)
    destination = normalize_locode_scac(destination)
    carrier = normalize_locode_scac(carrier)

    filt = ScheduleFilter(
        origin=origin,
        destination=destination,
        date_from=date_from,
        date_to=date_to,
        equipment=equipment,
        routingType=routingType,
        carrier=carrier,
        sort=sort,
        nContainers=nContainers,
    )

    page_params = Page(page=page, pageSize=pageSize)

    try:
        items, meta = provider.list(filt, page_params)
    except SearoutesRateLimitError as e:
        raise HTTPException(status_code=429, detail=e.to_dict())
    except SearoutesAPIError as e:
        # Map to appropriate HTTP status or default to 502 for upstream errors
        status_code = 502 if e.code and "HTTP_5" in e.code else 400
        raise HTTPException(status_code=status_code, detail=e.to_dict())
    except SearoutesError as e:
        raise HTTPException(status_code=502, detail=e.to_dict())

    # Return the exact envelope format specified in the contract
    return {
        "items": [item.model_dump() for item in items],
        "total": meta.total,
        "page": meta.page,
        "pageSize": meta.pageSize,
    }


@router.get("/api/schedules.csv")
def export_schedules_csv(
    origin: Optional[str] = None,
    destination: Optional[str] = None,
    date_from: Optional[str] = Query(None, alias="from"),
    date_to: Optional[str] = Query(None, alias="to"),
    equipment: Optional[str] = None,
    routingType: Optional[str] = None,
    carrier: Optional[str] = None,
    sort: str = "etd",
):
    """
    Export schedules as CSV with filtering and sorting.

    This endpoint uses the same filters and sorting as the main schedules endpoint,
    but ignores pagination and returns all matching results in CSV format.

    CSV columns (in order):
    originLocode,destinationLocode,etd,eta,vessel,voyage,carrier,routingType,transitDays,service
    """
    # Use same filter logic as main endpoint
    filt = ScheduleFilter(
        origin=origin,
        destination=destination,
        date_from=date_from,
        date_to=date_to,
        equipment=equipment,
        routingType=routingType,
        carrier=carrier,
        sort=sort,
    )

    # Get ALL results (ignore pagination for exports)
    page_params = Page(page=1, pageSize=10000)  # Large page size to get all results
    items, meta = provider.list(filt, page_params)

    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header row with exact column order from spec
    writer.writerow(
        [
            "originLocode",
            "destinationLocode",
            "etd",
            "eta",
            "vessel",
            "voyage",
            "carrier",
            "routingType",
            "transitDays",
            "service",
        ]
    )

    # Write data rows
    for item in items:
        writer.writerow(
            [
                extract_locode(item.origin),
                extract_locode(item.destination),
                item.etd,
                item.eta,
                item.vessel,
                item.voyage,
                item.carrier,
                item.routingType,
                item.transitDays,
                item.service or "",  # Handle None values
            ]
        )

    csv_content = output.getvalue()
    output.close()

    # Return CSV with proper headers
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=schedules.csv"},
    )


@router.get("/api/schedules.xlsx")
def export_schedules_xlsx(
    origin: Optional[str] = None,
    destination: Optional[str] = None,
    date_from: Optional[str] = Query(None, alias="from"),
    date_to: Optional[str] = Query(None, alias="to"),
    equipment: Optional[str] = None,
    routingType: Optional[str] = None,
    carrier: Optional[str] = None,
    sort: str = "etd",
):
    """
    Export schedules as Excel (.xlsx) with filtering and sorting.

    This endpoint uses the same filters and sorting as the main schedules endpoint,
    but ignores pagination and returns all matching results in Excel format.

    Creates a workbook with a "Schedules" sheet containing the same columns as CSV:
    originLocode,destinationLocode,etd,eta,vessel,voyage,carrier,routingType,transitDays,service
    """
    # Use same filter logic as main endpoint
    filt = ScheduleFilter(
        origin=origin,
        destination=destination,
        date_from=date_from,
        date_to=date_to,
        equipment=equipment,
        routingType=routingType,
        carrier=carrier,
        sort=sort,
    )

    # Get ALL results (ignore pagination for exports)
    page_params = Page(page=1, pageSize=10000)  # Large page size to get all results
    items, meta = provider.list(filt, page_params)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedules"

    # Define headers with exact column order from spec
    headers = [
        "originLocode",
        "destinationLocode",
        "etd",
        "eta",
        "vessel",
        "voyage",
        "carrier",
        "routingType",
        "transitDays",
        "service",
    ]

    # Write header row
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    for row_idx, item in enumerate(items, 2):  # Start from row 2 (after header)
        ws.cell(row=row_idx, column=1, value=extract_locode(item.origin))
        ws.cell(row=row_idx, column=2, value=extract_locode(item.destination))
        ws.cell(row=row_idx, column=3, value=item.etd)
        ws.cell(row=row_idx, column=4, value=item.eta)
        ws.cell(row=row_idx, column=5, value=item.vessel)
        ws.cell(row=row_idx, column=6, value=item.voyage)
        ws.cell(row=row_idx, column=7, value=item.carrier)
        ws.cell(row=row_idx, column=8, value=item.routingType)
        ws.cell(row=row_idx, column=9, value=item.transitDays)
        ws.cell(row=row_idx, column=10, value=item.service or "")

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].auto_size = True

    # Save to BytesIO buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    excel_content = output.getvalue()
    output.close()

    # Return Excel with proper headers
    return Response(
        content=excel_content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=schedules.xlsx"},
    )


@router.get("/api/schedules/{hash}/co2")
def get_co2_details(hash: str):
    """
    Get CO₂ details for a specific itinerary hash.

    Proxies to Searoutes CO₂ API endpoint for detailed emissions information.

    Args:
        hash: Searoutes itinerary hash

    Returns:
        CO₂ details from Searoutes API

    Raises:
        HTTPException: If hash not found or API error
    """
    # Check if provider supports CO2 details (only Searoutes provider does)
    if not hasattr(provider, "_make_request"):
        raise HTTPException(status_code=501, detail="CO₂ details not supported by current provider")

    try:
        # Proxy request to Searoutes CO₂ API endpoint
        response = provider._make_request(f"/co2/v2/execution/{hash}")
        co2_data = response.json()

        return co2_data

    except Exception as e:
        # Handle various error types from Searoutes API
        if hasattr(e, "response") and e.response:
            status_code = e.response.status_code
            if status_code == 404:
                raise HTTPException(
                    status_code=404, detail=f"CO₂ details not found for hash: {hash}"
                )
            elif status_code >= 400:
                raise HTTPException(
                    status_code=status_code, detail=f"Searoutes API error: {str(e)}"
                )

        raise HTTPException(status_code=500, detail=f"Error retrieving CO₂ details: {str(e)}")
