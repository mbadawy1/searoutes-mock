from fastapi import FastAPI, Query, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from typing import Optional, List, Dict, Any
import json, os, pathlib, datetime, io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

app = FastAPI(title="Searoutes Schedules Mock", version="0.2")

# CORS (allow everything for local dev)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

ROOT = pathlib.Path(__file__).parent
FIX = ROOT / "fixtures"
DATA = ROOT / "data"

# Serve the viewer & static files from /ui (so you can open http://127.0.0.1:4010/ui/viewer.html)
app.mount("/ui", StaticFiles(directory=str(ROOT), html=True), name="ui")

@app.get("/")
def root():
    return RedirectResponse(url="/ui/viewer.html")

def load_fixture(from_locode: str, to_locode: str, equipment: Optional[str]) -> dict:
    # Try exact match with equipment first, then common fallbacks
    candidates = []
    if equipment:
        candidates.append(f"{from_locode.upper()}-{to_locode.upper()}-{equipment.upper()}.json")
    candidates.append(f"{from_locode.upper()}-{to_locode.upper()}-40HC.json")
    candidates.append(f"{from_locode.upper()}-{to_locode.upper()}-40RF.json")
    for name in candidates:
        fpath = FIX / name
        if fpath.exists():
            return json.loads(fpath.read_text())
    return {"items": []}

def parse_dt(s: str | None):
    if not s: return None
    try:
        return datetime.datetime.fromisoformat(s)
    except Exception:
        try:
            # Try forgiving parse (strip Z if present)
            return datetime.datetime.fromisoformat(s.replace('Z',''))
        except Exception:
            return None

def filter_items(items: List[Dict[str, Any]], carrierScac: Optional[str], fromDate: Optional[str], toDate: Optional[str]) -> List[Dict[str, Any]]:
    # filter by carrier
    if carrierScac:
        c = carrierScac.upper()
        filt = []
        for it in items:
            feats = [f for f in it.get("features", []) if f.get("properties", {}).get("carrierScac","").upper() == c]
            if feats:
                filt.append({"hash": it.get("hash"), "features": feats})
        items = filt
    # filter by window (first leg departure)
    if fromDate or toDate:
        start = datetime.datetime.fromisoformat(fromDate) if fromDate else None
        end   = datetime.datetime.fromisoformat(toDate) if toDate else None
        win = []
        for it in items:
            if not it.get("features"): 
                continue
            first_dep = parse_dt(it["features"][0]["properties"]["departure"]["time"])
            if first_dep is None: 
                continue
            if start and first_dep < start: 
                continue
            if end and first_dep > end: 
                continue
            win.append(it)
        items = win
    return items

@app.get("/itinerary/v2/execution")
def execution(
    fromLocode: str = Query(...),
    toLocode: str = Query(...),
    carrierScac: Optional[str] = Query(None),
    fromDate: Optional[str] = Query(None),
    toDate: Optional[str] = Query(None),
    equipment: Optional[str] = Query(None),
    sortBy: Optional[str] = Query(None)
):
    data = load_fixture(fromLocode, toLocode, equipment)
    items = filter_items(data.get("items", []), carrierScac, fromDate, toDate)

    # Optional sort by total transit time
    if sortBy and sortBy.upper() == "TRANSIT_TIME":
        def total_transit(it):
            days = 0
            for f in it.get("features", []):
                days += f["properties"].get("transitTimeDays", 0)
            return days
        items = sorted(items, key=total_transit)

    return JSONResponse({"items": items})

# ----- Search helpers (simulate provider helpers) -----
@app.get("/ports/search")
def ports_search(q: str = Query(..., description="name, alias, or LOCODE"), country: Optional[str] = None, limit: int = 15):
    pth = DATA / "ports.json"
    if not pth.exists(): return {"items": []}
    ports = json.loads(pth.read_text())

    ql = q.strip().lower()
    cc = country.upper() if country else None
    scored = []
    for p in ports:
        if cc and p.get("country","").upper() != cc:
            continue
        hay = [p.get("name",""), p.get("locode",""), p.get("countryName","")] + p.get("aliases",[])
        hay = [h.lower() for h in hay]
        if any(ql in h for h in hay):
            score = 0
            name = p.get("name","").lower()
            if ql in name: score -= 10
            if ql == p.get("locode","").lower(): score -= 20
            scored.append((score, p))
    scored.sort(key=lambda x: x[0])
    return {"items": [s[1] for s in scored[:limit]]}

@app.get("/carriers/search")
def carriers_search(q: Optional[str] = None, limit: int = 15):
    pth = DATA / "carriers.json"
    if not pth.exists(): return {"items": []}
    carr = json.loads(pth.read_text())
    if not q:
        return {"items": carr[:limit]}
    ql = q.strip().lower()
    out = [c for c in carr if ql in c.get("name","").lower() or ql in c.get("scac","").lower()]
    return {"items": out[:limit]}

# ----- Export helpers -----
def map_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows = []
    for it in items:
        feats = it.get("features", [])
        if not feats: 
            continue
        legs = []
        for i, f in enumerate(feats, start=1):
            p = f.get("properties", {})
            legs.append({
                "seq": i,
                "fromLocode": p.get("departure", {}).get("locode"),
                "fromTime": p.get("departure", {}).get("time"),
                "toLocode": p.get("arrival", {}).get("locode"),
                "toTime": p.get("arrival", {}).get("time"),
                "carrier": p.get("carrier"),
                "scac": p.get("carrierScac"),
                "service": p.get("serviceId"),
                "vessel": (p.get("vessel") or {}).get("name"),
                "imo": (p.get("vessel") or {}).get("imo"),
                "voyage": (p.get("vessel") or {}).get("voyage"),
                "transitDays": p.get("transitTimeDays"),
            })
        first = legs[0]
        last  = legs[-1]
        routing = "Direct" if len(legs) <= 1 else f"Transshipment ×{len(legs)-1}"
        total_transit = sum(l.get("transitDays") or 0 for l in legs)
        rows.append({
            "DepartureLocode": first["fromLocode"],
            "ETD": first["fromTime"],
            "ArrivalLocode": last["toLocode"],
            "ETA": last["toTime"],
            "Carrier": first["carrier"],
            "SCAC": first["scac"],
            "Service": first["service"],
            "Vessel": first["vessel"],
            "Voyage": first["voyage"],
            "IMO": first["imo"],
            "TransitDays": total_transit,
            "LegsCount": len(legs),
            "RoutingType": routing,
        })
    return rows

@app.get("/export/csv")
def export_csv(fromLocode: str, toLocode: str, carrierScac: Optional[str] = None, fromDate: Optional[str] = None, toDate: Optional[str] = None, equipment: Optional[str] = None):
    data = load_fixture(fromLocode, toLocode, equipment)
    items = filter_items(data.get("items", []), carrierScac, fromDate, toDate)
    rows = map_rows(items)
    headers = ["DepartureLocode","ETD","ArrivalLocode","ETA","Carrier","SCAC","Service","Vessel","Voyage","IMO","TransitDays","LegsCount","RoutingType"]
    lines = [",".join(headers)]
    def esc(v): 
        s = "" if v is None else str(v)
        return '"' + s.replace('"','""') + '"'
    for r in rows:
        lines.append(",".join(esc(r[h]) for h in headers))
    import io
    csv_bytes = ("\n".join(lines)).encode("utf-8")
    fname = f"schedules_{fromLocode}_{toLocode}.csv"
    return StreamingResponse(io.BytesIO(csv_bytes), media_type="text/csv", headers={"Content-Disposition": f'attachment; filename=\"{fname}\"'})

@app.get("/export/xlsx")
def export_xlsx(fromLocode: str, toLocode: str, carrierScac: Optional[str] = None, fromDate: Optional[str] = None, toDate: Optional[str] = None, equipment: Optional[str] = None):
    data = load_fixture(fromLocode, toLocode, equipment)
    items = filter_items(data.get("items", []), carrierScac, fromDate, toDate)
    rows = map_rows(items)

    wb = Workbook(); ws = wb.active; ws.title = "Schedules"
    headers = ["DepartureLocode","ETD","ArrivalLocode","ETA","Carrier","SCAC","Service","Vessel","Voyage","IMO","TransitDays","LegsCount","RoutingType"]
    ws.append(headers)
    for r in rows: ws.append([r.get(h) for h in headers])
    widths = [16,20,16,20,20,8,10,22,10,12,12,10,16]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    import io
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    fname = f"schedules_{fromLocode}_{toLocode}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename=\"{fname}\"'})
